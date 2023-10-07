import multiprocessing
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import MySQLdb
import requests
from time import sleep
from datetime import datetime
from pathlib import Path
import openpyxl
import re

mode1 = 0
mode2 = 0
mode3 = 0
mode4 = 0

time = datetime.now()
timestr = time.strftime("%Y%m%d_%H%M")

def find_element_by_css(driver, css_selector):
    try:
        return driver.find_element(By.CSS_SELECTOR, css_selector)
    except NoSuchElementException as _:
        return None
    
def saleper_calc(price, saleprice):
    price = int(price.strip('₩ ').replace(',', ''))
    saleprice = int(saleprice.strip('₩ ').replace(',', ''))
    _ = price - saleprice
    __ = int((_ / price) * 100)
    saleper = "-"+str(__)+"%"
    return saleper

def Driver_Start(platform, URL):
    #크롬드라이버 옵션 설정
    options = Options()
    options.add_experimental_option("detach", True)
    options.add_argument("headless")
    options.add_argument("disable-gpu")
    options.add_argument("lang=ko_KR")
    options.add_argument('user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36')

    #크롬드라이버 인스턴스 생성 및 옵션, 크기, URL, 암시적 대기 설정
    driver = webdriver.Chrome(options=options)
    driver.implicitly_wait(10)
    driver.set_window_size(1920,1080)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": """ Object.defineProperty(navigator, 'webdriver', { get: () => undefined }) """})
    driver.execute_script("Object.defineProperty(navigator, 'plugins', {get: function() {return[1, 2, 3, 4, 5];},});")
    driver.execute_script("Object.defineProperty(navigator, 'languages', {get: function() {return ['ko-KR', 'ko']}})")
    driver.get(URL)

    print(platform,"driver 생성 완료")
    return driver

def nintendo_crawling():
    platform = 'switch'
    excel_gamedata = openpyxl.Workbook()
    excel_gamedata_sheet = excel_gamedata.active
    gamedata_row_column = ["TITLE", "PRICE", "SALEPRICE", "SALEPER", "DESCRIPTION", "IMGDATA", "GAMEIMG", "URL"]
    excel_gamedata_sheet.append(gamedata_row_column)
    sleep(1)
    excel_gamegenre = openpyxl.Workbook()
    excel_gamegenre_sheet = excel_gamegenre.active
    gamegenre_row_column = ["TITLE", "GENRE"]
    excel_gamegenre_sheet.append(gamegenre_row_column)
    sleep(1)
    path = f'C:\Crawling_Excel_File\{timestr}'
    Path(path).mkdir(parents=True, exist_ok=True)
    
    driver = Driver_Start(platform, 'https://store.nintendo.co.kr/games/sale')
    sleep(1)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    error = soup.select_one('h1')
    if error.text == '403 Forbidden':
        print("4분30초간 일시정지(403 forbidden 오류 회피를 위해)")
        driver.quit()
        sleep(270)
        print("재실행")
        driver = Driver_Start(platform, 'https://store.nintendo.co.kr/games/sale')

    sleep(5)
    driver.find_element(By.CLASS_NAME, 'popup-close').click()
    sleep(2)
    page = driver.find_element(By.TAG_NAME, "body")
    for _ in range(0,130):
        page.send_keys(Keys.PAGE_DOWN)
        sleep(0.3)

    soup = BeautifulSoup(driver.page_source, "html.parser")

    #게임 데이터 크롤링
    panel = soup.select_one("div.category-product-list")
    gamelist = panel.select("div.category-product-item")

    i=0
    #l=0
    for item in gamelist:
        title = item.find("a", class_="category-product-item-title-link").text.strip()
        price = item.find("span", attrs={"data-price-type" : "oldPrice"})
        saleprice = item.find("span", attrs={"data-price-type" : "finalPrice"})
        img = item.find("img", class_="product-image-photo mplazyload mplazyload-transparent entered loaded")
        if img != None:
            imgdata = img["src"]

        if price == None:
            price = item.find("span", class_="point-icon-wrapper")
            price = price.text + "포인트"
            saleprice = "포인트로 구매 가능"
            saleper = "X"
        else:
            price = price.find("span").text
            saleprice = saleprice.find("span").text
            saleper = saleper_calc(price, saleprice)

        game_link = item.select_one('a')
        move = game_link["href"]

        driver.execute_script(f'window.open(\'{move}\');')
        driver.switch_to.window(driver.window_handles[-1])
        sleep(1.5)
        new_soup = BeautifulSoup(driver.page_source, "html.parser")

        error = new_soup.select_one('h1')
        if error.text == '403 Forbidden':
            print("4분30초간 일시정지(403 forbidden 오류 회피를 위해)")
            driver.close()
            driver.switch_to.window(driver.window_handles[-1])
            sleep(270)
            print("재실행")
            driver.execute_script(f'window.open(\'{move}\');')
            driver.switch_to.window(driver.window_handles[-1])
            sleep(5)
            new_soup = BeautifulSoup(driver.page_source, "html.parser")


        descript = new_soup.select_one("div.game_ex")
        description = ""
        if descript == None:
            descript = new_soup.select_one("div.value")
        else:
            descript = None
        
        if descript != None:
            descript = descript.text.strip()
            for char in descript:
                if char.isalnum() or char.isspace():
                    description += char

        gameimg = new_soup.select_one('img.fotorama__img')
        if gameimg == None:
            print("gameimg 오류 3초 대기후 재크롤링")
            sleep(3)
            new_soup = BeautifulSoup(driver.page_source, "html.parser")
            gameimg = new_soup.select_one('img.fotorama__img') 

        if gameimg != None:
            gameimg = gameimg["src"]

        gamedata_column = [title, price, saleprice, saleper, description, imgdata, gameimg, move]
        excel_gamedata_sheet.append(gamedata_column)

        tagparent = new_soup.find("div", class_="product-attribute game_category")
        if tagparent != None:
            tag = tagparent.find("div", class_="product-attribute-val").text.split(',')
            tag_length = len(tag)
            num = 0

            while num < tag_length:
                tag[num] = tag[num].strip()
                gamegenre_column = [title, tag[num]]
                excel_gamegenre_sheet.append(gamegenre_column)
                num += 1
        
        i+=1
        #l+=1

        print(f'switch - {i}.{title} 입력 완료')
        driver.close()
        driver.switch_to.window(driver.window_handles[-1])

    newtime = datetime.now()
    newtimestr = newtime.strftime("%Y%m%d_%H%M")

    sleep(2)
    excel_gamedata.save(f'{path}\Switch_Crawling.xlsx')
    excel_gamegenre.save(f'{path}\Switch_Genre_Crawling.xlsx')
    excel_gamedata.close()
    excel_gamegenre.close()
    print(platform, " 크롤링 완료 시작시간:", timestr, ", 완료시간:", newtimestr)
    mode1 = 1
    driver.quit()

def ps_crawling():
    platform = 'ps'
    gameURL = 'https://store.playstation.com/'
    excel_gamedata = openpyxl.Workbook()
    excel_gamedata_sheet = excel_gamedata.active
    gamedata_row_column = ["TITLE", "PRICE", "SALEPRICE", "SALEPER", "DESCRIPTION", "IMGDATA", "GAMEIMG", "URL"]
    excel_gamedata_sheet.append(gamedata_row_column)
    sleep(1)
    excel_gamegenre = openpyxl.Workbook()
    excel_gamegenre_sheet = excel_gamegenre.active
    gamegenre_row_column = ["TITLE", "GENRE"]
    excel_gamegenre_sheet.append(gamegenre_row_column)
    sleep(1)
    path = f'C:\Crawling_Excel_File\{timestr}'
    Path(path).mkdir(parents=True, exist_ok=True)

    driver = Driver_Start(platform, 'https://store.playstation.com/ko-kr/pages/deals')

    sleep(2)
    driver.find_element(By.XPATH, "//*[@id='main']/div/div[3]/section/div/ul/li[3]/a").click()
    sleep(2)

    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")

    #할인 페이지 들어오자마자 페이지바 읽어와서 마지막 페이지 값 저장
    pagebar = driver.find_element(By.CLASS_NAME, "psw-l-stack-center")
    pages = pagebar.find_elements(By.CSS_SELECTOR, 'li')

    #다음페이지 초기값(숫자)
    next_page = 2
    #마지막페이지 값(문자열)
    last_page = pages[4].text
    #마지막 게임 이름 초기값(마지막페이지에 game_list가 한개 존재할경우 그 한개를 반복해서 저장하길래 그런상황에서 break를 걸기위해 변수 하나 생성)
    last_game_title = ""
    #db에 넣을때 title이 중복되면 오류나서 중복되는 게임은 스킵하기위해 beforetitle 선언
    beforetitle = ""

    restart = 1

    for roof in range(1000000):
        if(restart > 4):
            new_url = driver.current_url
            driver.quit()
            driver = Driver_Start(platform, new_url)
            print("셀레니움 재실행(렉방지)")
            sleep(3)
            restart = 1

        #페이지가 변경됐을때 변경된 페이지를 파싱
        html = driver.page_source
        soup = BeautifulSoup(html, "html.parser")

        sleep(1.5)

        panel = soup.find("div", class_="psw-l-w-1/1")
        game_list = panel.find_all("li", class_="psw-l-w-1/2@mobile-s psw-l-w-1/2@mobile-l psw-l-w-1/6@tablet-l psw-l-w-1/4@tablet-s psw-l-w-1/6@laptop psw-l-w-1/8@desktop psw-l-w-1/8@max")

        for game in game_list:
            try:
                title = game.find("span", class_="psw-t-body psw-c-t-1 psw-t-truncate-2 psw-m-b-2").text
            except:
                _ = 0
                while(title == None):
                    if _ > 5:
                        print("title 크롤링 오류 재실행할것(트래픽이 느린것으로 추정)")
                        driver.quit()
                        break
                    sleep(3)
                    title = game.find("span", class_="psw-t-body psw-c-t-1 psw-t-truncate-2 psw-m-b-2").text
                    _  += 1

            if title == beforetitle:
                continue
            #last_game_title의 값과 현재 크롤링하는 게임의 title의 값이 같으면 크롤링 종료
            if last_game_title == title:
                break

            img = game.find("img", class_="psw-fade-in psw-top-left psw-l-fit-cover")

            try:
                imgdata = img["src"]
            except:
                _ = 0
                while(imgdata == None):
                    if _ > 5:
                        print("imgdata 크롤링 오류 재실행할것(트래픽이 느린것으로 추정)")
                        driver.quit()
                        break
                    sleep(3)
                    imgdata = game.find("img", class_="psw-fade-in psw-top-left psw-l-fit-cover")["src"]
                    _  += 1
            
            saleper = game.find("span", class_="psw-body-2 psw-badge__text psw-badge--none psw-text-bold psw-p-y-0 psw-p-2 psw-r-1 psw-l-anchor")

            #64라인 saleper가 None이면 할인율은 무조건 -100%기때문에 price, saleprice = 무료
            if saleper == None:
                saleper = game.find("span", class_="psw-body-2 psw-badge__text psw-badge--ps-plus psw-c-bg-ps-plus psw-text-bold psw-p-y-0 psw-p-2 psw-r-1 psw-l-anchor")
                price = "무료"
                saleprice = "무료"
                #-100% 할인도 적혀있지않다면 구매할수없는 게임이기에 정가, 할인가, 할인율의 값을 "구매할 수 없음"으로 설정
                if saleper == None:
                    price = game.find("span", class_="psw-m-r-3").text
                    saleprice = "X"
                    saleper = "X"
                else:
                    saleper = saleper.text
            #saleper가 None이 아니였기때문에 price와 saleprice를 크롤링
            else:
                price = game.find("s").text
                saleprice = game.find("span", class_="psw-m-r-3").text
                saleper = saleper.text

            #현재 페이지와 마지막 페이지의 숫자가 똑같다면 last_game_title에 방금 크롤링한 게임 이름을 저장함
            if next_page-1 == int(last_page):
                last_game_title = title

            game_link = game.select_one('a')["href"]
            move = gameURL + game_link

            driver.execute_script(f'window.open(\'{move}\');')
            driver.switch_to.window(driver.window_handles[-1])
            sleep(1.5)
            new_soup = BeautifulSoup(driver.page_source, "html.parser")
            sleep(1.5)

            descript = new_soup.find('p', attrs={'data-qa': 'mfe-game-overview#description'})

            if descript != None:
                descript = descript.text.strip()
                description = ""
                for char in descript:
                    if char.isalnum() or char.isspace():
                        description += char

            try:
                gameimg = new_soup.find('img', attrs={'data-qa': 'gameBackgroundImage#heroImage#image'})
                if gameimg == None:
                    gameimg = new_soup.find('img', attrs={'data-qa': 'gameBackgroundImage#tileImage#image'})
            except TypeError as e:
                print(e)
                _ = 0
                while(gameimg==None):
                    if _ > 5:
                        print("gameimg 크롤링 오류 재실행할것(트래픽이 느린것으로 추정)")
                        driver.quit()
                        break
                    sleep(3)
                    print("gameimg 추출중")
                    gameimg = new_soup.find('img', attrs={'data-qa': 'gameBackgroundImage#heroImage#image'})
                    if gameimg == None:
                        gameimg = new_soup.find('img', attrs={'data-qa': 'gameBackgroundImage#tileImage#image'})
                    _ += 1
            except:
                gameimg = None

            if gameimg != None:
                gameimg = gameimg["src"]
            
            sleep(0.5)
            
            gamedata_column = [title, price, saleprice, saleper, description, imgdata, gameimg, move]
            excel_gamedata_sheet.append(gamedata_column)
            
            tagparent = new_soup.find("dd", {'data-qa':'gameInfo#releaseInformation#genre-value'})
            if tagparent != None:
                tag = tagparent.select_one("span").text.split(',')
                tag_length = len(tag)
                num = 0
            
            while num < tag_length:
                tag[num] = tag[num].strip()
                gamegenre_column = [title, tag[num]]
                excel_gamegenre_sheet.append(gamegenre_column)
                num += 1

            beforetitle = title

            print(f'ps - {next_page-1}p.{title} 입력 완료')
            driver.close()
            driver.switch_to.window(driver.window_handles[-1])

        #마지막페이지 크롤링 끝나면 break로 while문 빠져나옴
        if int(last_page)+1 == next_page:
            break
        
        sleep(1.5)

        #페이지 이동 실행하는 부분
        pagebar = driver.find_element(By.CLASS_NAME, "psw-l-stack-center")
        pages = pagebar.find_elements(By.CSS_SELECTOR, 'li')
        #화면 하단에 페이지 이동바의 숫자(2,3,4,...)와 next_page 변수값이 같다면 다음페이지가 존재하는것이니 다음페이지로 이동
        for page in pages:
            if page.text == str(next_page):
                page.click()
                next_page += 1
                restart += 1
                sleep(1.5)
                break

    newtime = datetime.now()
    newtimestr = newtime.strftime("%Y%m%d_%H%M")

    excel_gamedata.save(f'{path}\PS_Crawling.xlsx')
    excel_gamegenre.save(f'{path}\PS_Genre_Crawling.xlsx')
    excel_gamedata.close()
    excel_gamegenre.close()

    sleep(2)
    print(platform, " 크롤링 완료 - 시작시간:", timestr, ", 완료시간:", newtimestr)
    mode2 = 1
    driver.quit()

def steam_crawling(driver, excel_gamedata_sheet, excel_gamegenre_sheet):
    driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")

    panel = soup.select_one("div.saleitembrowser_SaleItemBrowserContainer_2wLns")
    gamelist = panel.select("div.salepreviewwidgets_SaleItemBrowserRow_y9MSd")

    for item in gamelist:
        title = item.find("div", class_="salepreviewwidgets_StoreSaleWidgetTitle_3jI46 StoreSaleWidgetTitle").text
        price = item.find("div", class_="salepreviewwidgets_StoreOriginalPrice_1EKGZ")
        saleprice = item.find("div", class_="salepreviewwidgets_StoreSalePriceBox_Wh0L8")
        saleper = item.find("div", class_="salepreviewwidgets_StoreSaleDiscountBox_2fpFv")
        
        game_link = item.select_one('div.salepreviewwidgets_StoreSaleWidgetHalfLeft_2Va3O > a')
        move = game_link["href"]
        
        #게임은 할인하는데 확장팩이 할인안하는 경우 price와 saleper가 안적혀있어서 if문 작성
        if price == None:
            price = "할인 X"
            saleprice = "할인 X"
            saleper = "할인 X"
        else:
            price = price.text
            saleprice = saleprice.text
            saleper = saleper.text

        driver.execute_script(f'window.open(\'{move}\');')
        driver.switch_to.window(driver.window_handles[-1])
        sleep(1.5)
        new_soup = BeautifulSoup(driver.page_source, "html.parser")

        check = new_soup.select_one("div.agegate_btn_ctn")
        if check != None:
            driver.find_element(By.XPATH, '//*[@id="ageYear"]').click()
            driver.find_element(By.XPATH, '//*[@id="ageYear"]/option[88]').click()
            driver.find_element(By.XPATH, '//*[@id="view_product_page_btn"]/span').click()
            sleep(1.5)
            new_soup = BeautifulSoup(driver.page_source, "html.parser")
            sleep(1.5)

        imgdata = new_soup.select_one('img.game_header_image_full')

        if imgdata == None:
            driver.execute_script(f'window.open(\'{move}\');')
            driver.switch_to.window(driver.window_handles[-1])
            if find_element_by_css(driver, '#ageYear') != None:
                driver.find_element(By.XPATH, '//*[@id="ageYear"]').click()
                driver.find_element(By.XPATH, '//*[@id="ageYear"]/option[88]').click()
                driver.find_element(By.XPATH, '//*[@id="view_product_page_btn"]/span').click()
                sleep(1.5)
            new_html = driver.page_source
            new_soup = BeautifulSoup(new_html, "html.parser")
            sleep(1.5)
            imgdata = new_soup.select_one('img.game_header_image_full') 
            if imgdata == None:
                imgdata = new_soup.select_one('img.package_header')

        if imgdata != None:
            imgdata = imgdata["src"]

        gameimg = new_soup.select_one('a.highlight_screenshot_link')

        if gameimg != None:
            gameimg = gameimg['href']

        descript = new_soup.select_one("div.game_description_snippet")
        description = ""
        if descript != None:
            descript = descript.text.strip()
            for char in descript:
                if char.isalnum() or char.isspace():
                    description += char

        tag = new_soup.select("a.app_tag")
        
        tag_length = len(tag)
        num = 0

        gamedata_column = [title, price, saleprice, saleper, description, imgdata, gameimg, move]
        excel_gamedata_sheet.append(gamedata_column)

        while num < tag_length:
            tag[num] = tag[num].text.strip()
            gamegenre_column = [title, tag[num]]
            excel_gamegenre_sheet.append(gamegenre_column)
            num += 1

        print(f'steam - {title} 입력 완료')

        #탭 종료후 원래 탭(베스트게임 페이지)으로 이동
        driver.close()
        driver.switch_to.window(driver.window_handles[-1])


def steam_start():
    platform = 'steam'
    url = 'https://store.steampowered.com/specials/'
    excel_gamedata = openpyxl.Workbook()
    excel_gamedata_sheet = excel_gamedata.active
    gamedata_row_column = ["TITLE", "PRICE", "SALEPRICE", "SALEPER", "DESCRIPTION", "IMGDATA", "GAMEIMG", "URL"]
    excel_gamedata_sheet.append(gamedata_row_column)
    sleep(1)
    excel_gamegenre = openpyxl.Workbook()
    excel_gamegenre_sheet = excel_gamegenre.active
    gamegenre_row_column = ["TITLE", "GENRE"]
    excel_gamegenre_sheet.append(gamegenre_row_column)
    sleep(1)
    path = f'C:\Crawling_Excel_File\{timestr}'
    Path(path).mkdir(parents=True, exist_ok=True)

    btn_num = 1
    btn_check = 1
    mode = 0
    for _ in range(10000000000):
        driver = Driver_Start(platform, url)
        action = ActionChains(driver)
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        sleep(1.5)
        steam_language_change(driver, soup)
        sleep(1.5)
        driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")
        sleep(1.5)
        driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")

        if mode == 0:
            #반복문으로 할인중인 게임 목록에 '더 보기' 버튼이 존재하면 해당 버튼을 누르고 없어지면 반복문을 빠져나오게 작성
            for __ in range(10000000000):
                button = find_element_by_css(driver, '#SaleSection_13268 > div.partnersaledisplay_SaleSection_2NfLq.eventbbcodeparser_SaleSectionCtn_2Xrw_.SaleSectionForCustomCSS > div.saleitembrowser_SaleItemBrowserContainer_2wLns > div:nth-child(2) > div.facetedbrowse_FacetedBrowseInnerCtn_hWbTI > div > div.saleitembrowser_ShowContentsContainer_3IRkb > button')
                if btn_num < 20:
                    #if else문 : '더 보기' 버튼이 존재하면(None 타입이 아니라 WebElement 타입이 반환되면)
                    if button != None:
                        #화면을 '더 보기' 버튼이 있는곳으로 이동시킨 후 버튼을 클릭한다
                        action.move_to_element(button).click().perform()
                        print("steam 버튼 클릭 - ", btn_num)
                        btn_num += 1
                        btn_check += 1
                    else: #'더 보기' 버튼이 존재하지않으면 break로 반복문 빠져나옴
                        print("버튼 전부 클릭 완료 - ", btn_check, "번 실행")
                        steam_crawling(driver, excel_gamedata_sheet, excel_gamegenre_sheet)
                        mode = 1
                        break
                else:
                    steam_crawling(driver, excel_gamedata_sheet, excel_gamegenre_sheet)
                    action.move_to_element(button).click().perform()
                    sleep(1.5)
                    url = driver.current_url
                    btn_num = 1
                    break
        elif mode == 1:
            newtime = datetime.now()
            newtimestr = newtime.strftime("%Y%m%d_%H%M")
            sleep(2)
            print(platform, " 크롤링 완료 - 시작시간:", timestr, ", 완료시간:", newtimestr)
            break
        driver.quit()

    excel_gamedata.save(f'{path}\Steam_Crawling.xlsx')
    excel_gamegenre.save(f'{path}\Steam_Genre_Crawling.xlsx')
    excel_gamedata.close()
    excel_gamegenre.close()
    driver.quit()
    mode3 = 1

def steam_language_change(driver, soup):
    #정보를 한글로 크롤링하기위해 페이지 언어 변경
    steam_language = soup.find("span", {"class": "pulldown global_action_link"})    
    if steam_language.string != "언어":
        driver.find_element(By.CSS_SELECTOR, "#language_pulldown").click()
        driver.find_element(By.CSS_SELECTOR, "#language_dropdown > div > a:nth-child(4)").click()
        print("steam 언어설정 완료")

def epic_crawling():
    platform = "epicgames"
    URL = 'https://store.epicgames.com/ko/browse'
    epicgames = 'https://store.epicgames.com'
    
    excel_gamedata = openpyxl.Workbook()
    excel_gamedata_sheet = excel_gamedata.active
    gamedata_row_column = ["TITLE", "PRICE", "SALEPRICE", "SALEPER", "DESCRIPTION", "IMGDATA", "GAMEIMG", "URL"]
    excel_gamedata_sheet.append(gamedata_row_column)
    sleep(1)
    excel_gamegenre = openpyxl.Workbook()
    excel_gamegenre_sheet = excel_gamegenre.active
    gamegenre_row_column = ["TITLE", "GENRE"]
    excel_gamegenre_sheet.append(gamegenre_row_column)
    sleep(1)
    path = f'C:\Crawling_Excel_File\{timestr}'
    Path(path).mkdir(parents=True, exist_ok=True)

    driver = Driver_Start(platform, URL)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    sleep(3)

    #할인 페이지 들어오자마자 페이지바 읽어와서 마지막 페이지 값 저장
    pages = driver.find_elements(By.CLASS_NAME, "css-12lid1g")
    #다음페이지 초기값(숫자)
    next_page = 2
    #마지막페이지 값(문자열)
    last_page = pages[6].text
    last_game_title = "title"

    for _ in range(10000000000):
        soup = BeautifulSoup(driver.page_source, "html.parser")
        panel = soup.select_one("section.css-zjpm9r")
        gamelist = panel.select("li.css-lrwy1y")
        
        driver.quit()
        for item in gamelist:
            saleper = item.find("div", class_="css-b0xoos")
            if saleper == None:
                continue
            else:
                price = item.find("div", class_="css-4jky3p").text
                saleprice = item.find("span", class_="css-119zqif").text
                saleper = saleper.text

            title = item.find("div", class_="css-rgqwpc")
            if title == None:
                title = item.find("div",class_="css-8uhtka").text
            else:
                title = title.text

            #직전 크롤링한 게임의 타이틀과 현재 크롤링중인 게임의 타이틀이 동일하면 마지막페이지라서 break
            if last_game_title == title:
                break

            game_link = item.select_one('a.css-g3jcms')["href"]
            move = epicgames+game_link
            print(f'{title}로 이동')

            driver = Driver_Start(platform, move)
            sleep(3)
            new_soup = BeautifulSoup(driver.page_source, "html.parser")

            if new_soup.select_one('h1.css-1gty6cv') != None:
                driver.quit()
                continue

            try:
                descript = new_soup.select_one("div.css-1myreog")
                description = ""
                if descript != None:
                    descript = descript.text.strip()
                    for char in descript:
                        if char.isalnum() or char.isspace():
                            description += char
            except:
                print("게임 설명 크롤링 오류 NULL로 대체")

            try:
                imgdata = new_soup.select_one('img.css-7i770w')["src"]
            except:
                pass
            
            try:
                gameimg_bar = new_soup.select_one('ul.css-elmzlf')
                if gameimg_bar != None:
                    gameimg = gameimg_bar.select_one('div.css-1q03292 > img')["src"]
                else:
                    gameimg = new_soup.select_one('img.css-1bbjmcj')["src"]
            except:
                pass

            tag = new_soup.select("li.css-t8k7")
            tag_length = len(tag)
            num = 0

            gamedata_column = [title, price, saleprice, saleper, description, imgdata, gameimg, move]
            excel_gamedata_sheet.append(gamedata_column)

            while num < tag_length:
                tag[num] = tag[num].text.strip()
                gamegenre_column = [title, tag[num]]
                excel_gamegenre_sheet.append(gamegenre_column)
                num += 1

            print(f'epic - {title} 입력 완료')
            driver.quit()

            #현재 페이지가 마지막 페이지면 last_game_title에 지금 크롤링중인 게임 타이틀 집어넣기
            if next_page-1 == int(last_page):
                last_game_title = title
            sleep(1)
        
        if int(last_page)+1 == next_page:
            break

        #페이지 이동 실행하는 부분
        sleep(1)
        pagebar = soup.select_one('ul.css-zks4l')
        pages = pagebar.select('a.css-1ns6940')
        for page in pages:
            if page.text == str(next_page):
                move_url = page["href"]
                move = epicgames + move_url
                next_page += 1
                driver.quit()
                driver = Driver_Start(platform, move)
                sleep(5)
                break

    newtime = datetime.now()
    newtimestr = newtime.strftime("%Y%m%d_%H%M")
    sleep(2)
    print(platform, " 크롤링 완료 - 시작시간:", timestr, ", 완료시간:", newtimestr)

    excel_gamedata.save(f'{path}\Epic_Crawling.xlsx')
    excel_gamegenre.save(f'{path}\Epic_Genre_Crawling.xlsx')
    excel_gamedata.close()
    excel_gamegenre.close()

    print("크롤링 끝 driver 종료")
    driver.quit()
    mode4 = 1

def danawa():
    URL = 'https://prod.danawa.com/game/index.php'
    driver = Driver_Start("danawa", URL)

    excel_releasedata = openpyxl.Workbook()
    excel_sheet = excel_releasedata.active
    gamedata_row_column = ["DATE", "TITLE", "PLATFORM", "PRICE"]
    excel_sheet.append(gamedata_row_column)
    path = f'C:\Crawling_Excel_File\{timestr}'
    Path(path).mkdir(parents=True, exist_ok=True)

    action = ActionChains(driver)

    before_date = ""

    for _ in range(10000000000000):
        soup = BeautifulSoup(driver.page_source, "html.parser")
        gamelist = soup.select_one('div.upc_tbl_wrap').select('tr')

        for list in gamelist:
            date = list.select_one('td.date')
            if date != None:
                date = date.text.strip()
                date = date.replace('.', '-')
                date = re.sub(r'\(.\)', '', date)

            if date == "":
                date = before_date

            platform = list.select_one('td.type')

            if platform != None:
                platform = platform.text.strip()
                if 'XBOX' in platform:
                    continue

            title = list.select_one('a.tit_link')
            if title != None:
                if date == None:
                    date = before_date
                title = title.text.strip()

            price = list.select_one('em.num')

            if price == None:
                price = list.select_one('span.p_txt')

            if price != None:
                price = price.text.strip()
            
            before_date = date
            if date != None and platform != None and title != None and price != None:
                gamedata_column = [date, title, platform, price]
                excel_sheet.append(gamedata_column)
                print(f'danawa - {title} DB 입력 완료')

        next_page_btn = driver.find_element(By.XPATH, '//*[@id="#nav_edge_next"]')
        action.click(next_page_btn).perform()

        sleep(3)

        new_soup = BeautifulSoup(driver.page_source, "html.parser")
        x = new_soup.select_one('p.n_txt')
        if x != None:
            break

    excel_releasedata.save(f'{path}\Danawa_Crawling.xlsx')
    excel_releasedata.close()

    sleep(2)
    print("크롤링 완료")
    driver.quit()

if __name__ == "__main__":
    services = Service(executable_path=ChromeDriverManager().install())
    options = Options()
    options.add_argument("headless")
    driver = webdriver.Chrome(service=services, options=options)
    driver.get('https://www.google.com')
    driver.quit()
    print("Chromedriver Install 완료")
    sleep(1)
    # 멀티프로세싱을 사용하여 각 코드 블록을 별도의 프로세스에서 실행
    process1 = multiprocessing.Process(target=nintendo_crawling)
    process2 = multiprocessing.Process(target=ps_crawling)
    process3 = multiprocessing.Process(target=steam_start)
    process4 = multiprocessing.Process(target=epic_crawling)
    process5 = multiprocessing.Process(target=danawa)

    """
    try:
        process1.start()
    except:
        print("process1 시작 오류")
    sleep(3)
    try:
        process2.start()
    except:
        print("process2 시작 오류")
    sleep(3)
    try:
        process3.start()
    except:
        print("process3 시작 오류")
    sleep(3)
    try:
        process4.start()
    except:
        print("process4 시작 오류")
    sleep(3)
    """
    try:
        process5.start()
    except:
        print("process5 시작 오류")

    # 프로세스 종료 대기
    try:
        process1.join()
    except:
        pass
    try:
        process2.join()
    except:
        pass    
    try:
        process3.join()
    except:
        pass    
    try:
        process4.join()
    except:
        pass
    try:
        process5.join()
    except:
        pass

    quit()